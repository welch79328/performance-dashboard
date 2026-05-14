import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.work_order import WorkOrder
from app.models.kpi import (
    TeamKPI, EfficiencyMetrics, EfficiencyByType, AgingOrder, WeeklyTrend,
)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
GREEN_FILL = PatternFill(start_color="6BCB77", end_color="6BCB77", fill_type="solid")


def _write_header(ws, row: int, headers: list[str]) -> None:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _write_kpi_summary(wb: Workbook, team_kpi: TeamKPI, date_label: str) -> None:
    ws = wb.create_sheet("工作量摘要")

    ws.cell(row=1, column=1, value=f"績效週報 — {date_label}").font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="團隊 KPI 摘要").font = Font(bold=True, size=12)

    kpi_data = [
        ("總工單量", team_kpi.total_tasks),
        ("已結案", team_kpi.completed_tasks),
        ("進行中", team_kpi.in_progress_tasks),
        ("結案率", f"{team_kpi.close_rate}%"),
        ("平均處理天數", team_kpi.avg_processing_days),
    ]
    for i, (label, value) in enumerate(kpi_data):
        ws.cell(row=4 + i, column=1, value=label)
        ws.cell(row=4 + i, column=2, value=value)

    # Member workloads
    row = 11
    ws.cell(row=row, column=1, value="成員工作量").font = Font(bold=True, size=12)
    row += 1
    _write_header(ws, row, ["姓名", "PM", "開發", "測試", "合計", "在手量"])
    row += 1

    for wl in team_kpi.member_workloads:
        ws.cell(row=row, column=1, value=wl.user_name)
        ws.cell(row=row, column=2, value=getattr(wl, "pm_count", 0))
        ws.cell(row=row, column=3, value=getattr(wl, "dev_count", 0))
        ws.cell(row=row, column=4, value=getattr(wl, "test_count", 0))
        ws.cell(row=row, column=5, value=getattr(wl, "total_count", 0))
        ws.cell(row=row, column=6, value=getattr(wl, "in_progress_count", 0))
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12


def _write_efficiency(wb: Workbook, efficiency: EfficiencyMetrics,
                       by_type: list[EfficiencyByType]) -> None:
    ws = wb.create_sheet("流程效率")

    ws.cell(row=1, column=1, value="整體效率").font = Font(bold=True, size=12)
    metrics = [
        ("端到端平均天數", efficiency.avg_total_days),
        ("開發階段平均天數", efficiency.avg_dev_days),
        ("測試階段平均天數", efficiency.avg_test_days),
        ("結案率", f"{efficiency.close_rate}%"),
        ("卡關率", f"{efficiency.stalled_rate}%"),
    ]
    for i, (label, value) in enumerate(metrics):
        ws.cell(row=2 + i, column=1, value=label)
        ws.cell(row=2 + i, column=2, value=value)

    row = 9
    ws.cell(row=row, column=1, value="依類型分析").font = Font(bold=True, size=12)
    row += 1
    _write_header(ws, row, ["類型", "數量", "平均天數", "開發耗時", "測試耗時", "結案率"])
    row += 1

    for bt in by_type:
        ws.cell(row=row, column=1, value=bt.type_name)
        ws.cell(row=row, column=2, value=bt.count)
        ws.cell(row=row, column=3, value=bt.avg_total_days)
        ws.cell(row=row, column=4, value=bt.avg_dev_days)
        ws.cell(row=row, column=5, value=bt.avg_test_days)
        ws.cell(row=row, column=6, value=f"{bt.close_rate}%")
        row += 1

    ws.column_dimensions["A"].width = 20


def _write_aging(wb: Workbook, aging_orders: list[AgingOrder]) -> None:
    ws = wb.create_sheet("未結案老化")

    _write_header(ws, 1, ["工單ID", "名稱", "客戶", "開發者", "指派日期", "已開天數", "嚴重度"])

    severity_fills = {"red": RED_FILL, "yellow": YELLOW_FILL, "green": GREEN_FILL}

    for i, ao in enumerate(aging_orders, 2):
        ws.cell(row=i, column=1, value=ao.id)
        ws.cell(row=i, column=2, value=ao.name)
        ws.cell(row=i, column=3, value=ao.client)
        ws.cell(row=i, column=4, value=ao.developer)
        ws.cell(row=i, column=5, value=ao.assign_date.isoformat())
        ws.cell(row=i, column=6, value=ao.days_open)
        severity_cell = ws.cell(row=i, column=7, value=ao.severity)
        if ao.severity in severity_fills:
            severity_cell.fill = severity_fills[ao.severity]

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12


def _write_order_detail(wb: Workbook, orders: list[WorkOrder]) -> None:
    ws = wb.create_sheet("工單明細")

    _write_header(ws, 1, [
        "工單ID", "名稱", "客戶", "類型", "PM", "開發者", "測試者",
        "指派日期", "測試日期", "結案日期", "狀態", "處理天數",
    ])

    for i, o in enumerate(orders, 2):
        ws.cell(row=i, column=1, value=o.id)
        ws.cell(row=i, column=2, value=o.name)
        ws.cell(row=i, column=3, value=o.client)
        ws.cell(row=i, column=4, value=o.type)
        ws.cell(row=i, column=5, value=o.pm)
        ws.cell(row=i, column=6, value=o.developer)
        ws.cell(row=i, column=7, value=o.tester)
        ws.cell(row=i, column=8, value=o.assign_date.isoformat() if o.assign_date else "")
        ws.cell(row=i, column=9, value=o.test_assign_date.isoformat() if o.test_assign_date else "")
        ws.cell(row=i, column=10, value=o.close_date.isoformat() if o.close_date else "")
        ws.cell(row=i, column=11, value=o.status)
        ws.cell(row=i, column=12, value=o.total_days if o.total_days is not None else "")

    ws.column_dimensions["B"].width = 40


def _write_trends(wb: Workbook, trends: list[WeeklyTrend]) -> None:
    ws = wb.create_sheet("週趨勢")

    _write_header(ws, 1, ["週起始", "週結束", "工單數", "結案率", "平均天數", "Bug數", "異動數"])

    for i, t in enumerate(trends, 2):
        ws.cell(row=i, column=1, value=t.week_start)
        ws.cell(row=i, column=2, value=t.week_end)
        ws.cell(row=i, column=3, value=t.task_count)
        ws.cell(row=i, column=4, value=f"{t.close_rate}%")
        ws.cell(row=i, column=5, value=t.avg_days)
        ws.cell(row=i, column=6, value=t.bug_count)
        ws.cell(row=i, column=7, value=t.change_count)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14


def generate_weekly_report(
    team_kpi: TeamKPI,
    orders: list[WorkOrder],
    efficiency: EfficiencyMetrics,
    efficiency_by_type: list[EfficiencyByType],
    aging_orders: list[AgingOrder],
    date_range_label: str,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    _write_kpi_summary(wb, team_kpi, date_range_label)
    _write_efficiency(wb, efficiency, efficiency_by_type)
    _write_aging(wb, aging_orders)
    _write_order_detail(wb, orders)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_monthly_report(
    team_kpi: TeamKPI,
    orders: list[WorkOrder],
    efficiency: EfficiencyMetrics,
    efficiency_by_type: list[EfficiencyByType],
    aging_orders: list[AgingOrder],
    trends: list[WeeklyTrend],
    date_range_label: str,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    _write_kpi_summary(wb, team_kpi, date_range_label)
    _write_efficiency(wb, efficiency, efficiency_by_type)
    _write_trends(wb, trends)
    _write_aging(wb, aging_orders)
    _write_order_detail(wb, orders)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
