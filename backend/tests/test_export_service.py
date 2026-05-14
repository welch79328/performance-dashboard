import io
from datetime import date

import pytest
from openpyxl import load_workbook

from app.models.work_order import WorkOrder
from app.models.kpi import (
    TeamKPI, PersonWorkload, EfficiencyMetrics, EfficiencyByType,
    AgingOrder, WeeklyTrend,
)
from app.services.export_service import generate_weekly_report, generate_monthly_report


def _wo(id: str, dev: str = "Lenny", client: str = "JGB", type: str = "開發",
        assign: str = "2026-05-01", close: str | None = None) -> WorkOrder:
    return WorkOrder(
        id=id, name=f"WO-{id}", developer=dev, client=client, type=type,
        assign_date=date.fromisoformat(assign),
        close_date=date.fromisoformat(close) if close else None,
    )


@pytest.fixture
def sample_team_kpi():
    return TeamKPI(
        department="pm_rd",
        total_tasks=10,
        completed_tasks=7,
        in_progress_tasks=3,
        close_rate=70.0,
        avg_processing_days=2.5,
        member_workloads=[
            PersonWorkload(user_name="Lenny", pm_count=0, dev_count=5, test_count=0, total_count=5, in_progress_count=2),
            PersonWorkload(user_name="Abu", pm_count=0, dev_count=3, test_count=0, total_count=3, in_progress_count=1),
        ],
        type_distribution={"開發": 5, "臭蟲": 3, "異動": 2},
        client_distribution={"JGB": 6, "富喬": 4},
    )


@pytest.fixture
def sample_orders():
    return [
        _wo("1", close="2026-05-03"),
        _wo("2", dev="Abu", client="富喬", type="臭蟲", close="2026-05-05"),
        _wo("3"),  # in progress
    ]


@pytest.fixture
def sample_efficiency():
    return EfficiencyMetrics(
        avg_total_days=2.5, avg_dev_days=1.5, avg_test_days=1.0,
        close_rate=70.0, stalled_rate=10.0,
    )


@pytest.fixture
def sample_efficiency_by_type():
    return [
        EfficiencyByType(type_name="開發", count=5, avg_total_days=3.0, close_rate=80.0),
        EfficiencyByType(type_name="臭蟲", count=3, avg_total_days=2.0, close_rate=66.7),
        EfficiencyByType(type_name="異動", count=2, avg_total_days=0.5, close_rate=100.0),
    ]


@pytest.fixture
def sample_aging():
    return [
        AgingOrder(id="3", name="WO-3", client="JGB", developer="Lenny",
                   assign_date=date(2026, 5, 1), days_open=13, severity="red"),
    ]


@pytest.fixture
def sample_trends():
    return [
        WeeklyTrend(week_start="2026-05-05", week_end="2026-05-11", task_count=5, close_rate=60.0),
        WeeklyTrend(week_start="2026-05-12", week_end="2026-05-18", task_count=8, close_rate=75.0),
    ]


# === Weekly report ===

class TestWeeklyReport:
    def test_returns_valid_xlsx(self, sample_team_kpi, sample_orders, sample_efficiency,
                                sample_efficiency_by_type, sample_aging):
        buf = generate_weekly_report(
            team_kpi=sample_team_kpi,
            orders=sample_orders,
            efficiency=sample_efficiency,
            efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging,
            date_range_label="2026-05-05 ~ 2026-05-11",
        )
        assert isinstance(buf, bytes)
        assert len(buf) > 0
        # Should be valid xlsx
        wb = load_workbook(io.BytesIO(buf))
        assert len(wb.sheetnames) >= 4

    def test_has_kpi_summary_sheet(self, sample_team_kpi, sample_orders,
                                    sample_efficiency, sample_efficiency_by_type, sample_aging):
        buf = generate_weekly_report(
            team_kpi=sample_team_kpi, orders=sample_orders,
            efficiency=sample_efficiency, efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging, date_range_label="2026-05-05 ~ 2026-05-11",
        )
        wb = load_workbook(io.BytesIO(buf))
        assert "工作量摘要" in wb.sheetnames

    def test_has_efficiency_sheet(self, sample_team_kpi, sample_orders,
                                  sample_efficiency, sample_efficiency_by_type, sample_aging):
        buf = generate_weekly_report(
            team_kpi=sample_team_kpi, orders=sample_orders,
            efficiency=sample_efficiency, efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging, date_range_label="2026-05-05 ~ 2026-05-11",
        )
        wb = load_workbook(io.BytesIO(buf))
        assert "流程效率" in wb.sheetnames

    def test_has_aging_sheet(self, sample_team_kpi, sample_orders,
                              sample_efficiency, sample_efficiency_by_type, sample_aging):
        buf = generate_weekly_report(
            team_kpi=sample_team_kpi, orders=sample_orders,
            efficiency=sample_efficiency, efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging, date_range_label="2026-05-05 ~ 2026-05-11",
        )
        wb = load_workbook(io.BytesIO(buf))
        assert "未結案老化" in wb.sheetnames

    def test_has_detail_sheet(self, sample_team_kpi, sample_orders,
                               sample_efficiency, sample_efficiency_by_type, sample_aging):
        buf = generate_weekly_report(
            team_kpi=sample_team_kpi, orders=sample_orders,
            efficiency=sample_efficiency, efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging, date_range_label="2026-05-05 ~ 2026-05-11",
        )
        wb = load_workbook(io.BytesIO(buf))
        assert "工單明細" in wb.sheetnames

    def test_kpi_summary_data(self, sample_team_kpi, sample_orders,
                               sample_efficiency, sample_efficiency_by_type, sample_aging):
        buf = generate_weekly_report(
            team_kpi=sample_team_kpi, orders=sample_orders,
            efficiency=sample_efficiency, efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging, date_range_label="2026-05-05 ~ 2026-05-11",
        )
        wb = load_workbook(io.BytesIO(buf))
        ws = wb["工作量摘要"]
        # Row 1 should be title/header
        # Should contain member data
        all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "Lenny" in all_values


# === Monthly report ===

class TestMonthlyReport:
    def test_returns_valid_xlsx(self, sample_team_kpi, sample_orders, sample_efficiency,
                                sample_efficiency_by_type, sample_aging, sample_trends):
        buf = generate_monthly_report(
            team_kpi=sample_team_kpi,
            orders=sample_orders,
            efficiency=sample_efficiency,
            efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging,
            trends=sample_trends,
            date_range_label="2026-05",
        )
        assert isinstance(buf, bytes)
        wb = load_workbook(io.BytesIO(buf))
        assert len(wb.sheetnames) >= 5

    def test_has_trend_sheet(self, sample_team_kpi, sample_orders, sample_efficiency,
                              sample_efficiency_by_type, sample_aging, sample_trends):
        buf = generate_monthly_report(
            team_kpi=sample_team_kpi, orders=sample_orders,
            efficiency=sample_efficiency, efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging, trends=sample_trends,
            date_range_label="2026-05",
        )
        wb = load_workbook(io.BytesIO(buf))
        assert "週趨勢" in wb.sheetnames

    def test_trend_data(self, sample_team_kpi, sample_orders, sample_efficiency,
                         sample_efficiency_by_type, sample_aging, sample_trends):
        buf = generate_monthly_report(
            team_kpi=sample_team_kpi, orders=sample_orders,
            efficiency=sample_efficiency, efficiency_by_type=sample_efficiency_by_type,
            aging_orders=sample_aging, trends=sample_trends,
            date_range_label="2026-05",
        )
        wb = load_workbook(io.BytesIO(buf))
        ws = wb["週趨勢"]
        all_values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
        assert "2026-05-05" in all_values
